import tkinter as tk
from tkinter import ttk, messagebox
import os
from typing import Optional
from tkinterhtml import HtmlFrame
from docx import Document
from openpyxl import load_workbook
from pdf2docx import Converter
import sys

class PreviewWindow:
    """Fenêtre pour la prévisualisation des fichiers"""
    
    def __init__(self, root: tk.Toplevel, file: dict):
        self.root = root
        self.file = file
        self.root.title(f"Prévisualisation - {file['filename']}")
        self.root.geometry("800x600")
        
        # Centrer la fenêtre
        self.center_window()
        
        # Créer l'interface
        self.create_widgets()
        
        # Charger le contenu avec débogage
        print(f"Débogage: Initialisation de PreviewWindow pour {self.file['filename']}, chemin: {self.file['filepath']}")
        self.load_content()
    
    def center_window(self):
        """Centrer la fenêtre"""
        self.root.update_idletasks()
        width = 800
        height = 600
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Créer les widgets"""
        # En-tête
        header = tk.Frame(self.root, bg='#17a2b8', height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        title_label = tk.Label(
            header,
            text=f"📄 Prévisualisation - {self.file['filename']}",
            font=('Segoe UI', 14, 'bold'),
            bg='#17a2b8',
            fg='white'
        )
        title_label.pack(side=tk.LEFT, padx=20, pady=15)
        
        close_button = tk.Button(
            header,
            text="✖️ Fermer",
            font=('Segoe UI', 10),
            bg='#dc3545',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            command=self.root.destroy
        )
        close_button.pack(side=tk.RIGHT, padx=20)
        
        # Zone de contenu
        self.content_frame = tk.Frame(self.root, bg='white')
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbar
        self.scrollbar = ttk.Scrollbar(self.content_frame, orient=tk.VERTICAL)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def load_content(self):
        """Charger le contenu du fichier"""
        extension = self.file['filename'].rsplit('.', 1)[-1].lower() if '.' in self.file['filename'] else ''
        
        if not os.path.exists(self.file['filepath']):
            messagebox.showerror("Erreur", f"Le fichier {self.file['filename']} n'existe pas à l'emplacement : {self.file['filepath']}")
            print(f"Débogage: Fichier non trouvé - {self.file['filepath']}")
            return
        
        print(f"Débogage: Tentative de prévisualisation de {self.file['filename']} (extension: {extension})")
        try:
            if extension == 'pdf':
                self.preview_pdf()
            elif extension in ['docx', 'doc']:
                self.preview_docx()
            elif extension in ['xlsx', 'xls']:
                self.preview_xlsx()
            else:
                self.show_unsupported_message()
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de prévisualiser {self.file['filename']}:\n{str(e)}")
            print(f"Débogage: Exception dans load_content - {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preview_pdf(self):
        """Prévisualiser un fichier PDF"""
        print(f"Débogage: Conversion PDF - {self.file['filepath']}")
        try:
            temp_html = "temp_preview.html"
            cv = Converter(self.file['filepath'])
            cv.convert(temp_html, start=0, end=1)  # Première page
            cv.close()
            
            html_frame = HtmlFrame(
                self.content_frame,
                vertical_scrollbar=self.scrollbar
            )
            html_frame.pack(fill=tk.BOTH, expand=True)
            with open(temp_html, 'r', encoding='utf-8') as f:
                html_frame.set_content(f.read())
            os.remove(temp_html)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la conversion PDF: {str(e)}")
            print(f"Débogage: Erreur PDF - {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preview_docx(self):
        """Prévisualiser un fichier DOCX"""
        print(f"Débogage: Lecture DOCX - {self.file['filepath']}")
        try:
            doc = Document(self.file['filepath'])
            text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
            
            text_area = tk.Text(
                self.content_frame,
                font=('Segoe UI', 10),
                wrap=tk.WORD,
                yscrollcommand=self.scrollbar.set,
                state='normal'
            )
            text_area.pack(fill=tk.BOTH, expand=True)
            text_area.insert(tk.END, text)
            text_area.config(state='disabled')
            self.scrollbar.config(command=text_area.yview)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture DOCX: {str(e)}")
            print(f"Débogage: Erreur DOCX - {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preview_xlsx(self):
        """Prévisualiser un fichier XLSX"""
        print(f"Débogage: Lecture XLSX - {self.file['filepath']}")
        try:
            wb = load_workbook(self.file['filepath'], read_only=True)
            sheet = wb.active
            text = ""
            for row in sheet.iter_rows(max_row=20):
                row_text = "\t".join([str(cell.value or '') for cell in row])
                text += row_text + "\n"
            
            text_area = tk.Text(
                self.content_frame,
                font=('Segoe UI', 10),
                wrap=tk.WORD,
                yscrollcommand=self.scrollbar.set,
                state='normal'
            )
            text_area.pack(fill=tk.BOTH, expand=True)
            text_area.insert(tk.END, text)
            text_area.config(state='disabled')
            self.scrollbar.config(command=text_area.yview)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture XLSX: {str(e)}")
            print(f"Débogage: Erreur XLSX - {str(e)}")
            import traceback
            traceback.print_exc()
    
    def show_unsupported_message(self):
        """Afficher un message pour les types non supportés"""
        tk.Label(
            self.content_frame,
            text="Prévisualisation non disponible pour ce type de fichier",
            font=('Segoe UI', 12),
            fg='#dc3545',
            bg='white'
        ).pack(pady=20)