import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from typing import Dict
import os

class FilePreviewWindow:
    """Fenêtre de prévisualisation de fichiers"""
    
    def __init__(self, root: tk.Toplevel, file: Dict, file_handler):
        self.root = root
        self.file = file
        self.file_handler = file_handler
        
        self.root.title(f"Prévisualisation - {file['filename']}")
        self.root.geometry("900x700")
        
        # Centrer la fenêtre
        self.center_window()
        
        # Créer l'interface
        self.create_widgets()
        
        # Charger la prévisualisation
        self.load_preview()
    
    def center_window(self):
        """Centrer la fenêtre"""
        self.root.update_idletasks()
        width = 900
        height = 700
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Créer les widgets"""
        # En-tête
        header = tk.Frame(self.root, bg='#4facfe', height=70)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        # Titre avec icône
        extension = self.file['filename'].rsplit('.', 1)[-1].lower() if '.' in self.file['filename'] else ''
        icon = self.file_handler.get_file_icon(extension)
        
        title_label = tk.Label(
            header,
            text=f"{icon} {self.file['filename']}",
            font=('Segoe UI', 14, 'bold'),
            bg='#4facfe',
            fg='white'
        )
        title_label.pack(side=tk.LEFT, padx=20, pady=20)
        
        # Bouton fermer
        close_button = tk.Button(
            header,
            text="✖️ Fermer",
            font=('Segoe UI', 10),
            bg='#dc3545',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            command=self.root.destroy
        )
        close_button.pack(side=tk.RIGHT, padx=20)
        
        # Bouton "Ouvrir avec application par défaut"
        open_btn = tk.Button(
            header,
            text="🔗 Ouvrir avec application",
            font=('Segoe UI', 10),
            bg='#28a745',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            command=self.open_with_default_app
        )
        open_btn.pack(side=tk.RIGHT, padx=10)
        
        # Frame de contenu avec scrollbar
        self.content_frame = tk.Frame(self.root, bg='white')
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    def load_preview(self):
        """Charger la prévisualisation selon le type de fichier"""
        extension = self.file['filename'].rsplit('.', 1)[-1].lower() if '.' in self.file['filename'] else ''
        
        if extension == 'pdf':
            self.preview_pdf()
        elif extension in ['txt', 'csv', 'log']:
            self.preview_text()
        elif extension in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            self.preview_image()
        elif extension in ['docx', 'doc']:
            self.preview_docx()
        elif extension in ['xlsx', 'xls']:
            self.preview_xlsx()
        else:
            self.show_no_preview()
    
    def preview_pdf(self):
        """Prévisualiser un PDF"""
        try:
            import fitz  # PyMuPDF
            
            # Créer un canvas avec scrollbar
            canvas = tk.Canvas(self.content_frame, bg='#f0f0f0', highlightthickness=0)
            scrollbar = ttk.Scrollbar(self.content_frame, orient=tk.VERTICAL, command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor=tk.NW)
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Ouvrir le PDF
            doc = fitz.open(self.file['filepath'])
            
            # Afficher les premières pages (limite à 5 pour les performances)
            max_pages = min(5, len(doc))
            
            tk.Label(
                scrollable_frame,
                text=f"📄 Affichage des {max_pages} première(s) page(s) sur {len(doc)}",
                font=('Segoe UI', 10),
                bg='#f0f0f0',
                fg='#6c757d'
            ).pack(pady=10)
            
            for page_num in range(max_pages):
                page = doc[page_num]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # Zoom x2
                
                # Convertir en image PIL puis PhotoImage
                from PIL import Image, ImageTk
                import io
                
                img = Image.open(io.BytesIO(pix.tobytes()))
                photo = ImageTk.PhotoImage(img)
                
                # Frame pour chaque page
                page_frame = tk.Frame(scrollable_frame, bg='white', relief=tk.SOLID, bd=1)
                page_frame.pack(pady=10, padx=20)
                
                tk.Label(
                    page_frame,
                    text=f"Page {page_num + 1}",
                    font=('Segoe UI', 9, 'bold'),
                    bg='white'
                ).pack(pady=5)
                
                label = tk.Label(page_frame, image=photo, bg='white')
                label.image = photo  # Garder une référence
                label.pack(padx=10, pady=10)
            
            doc.close()
            
            # Scroll avec molette
            def on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", on_mousewheel)
            
        except ImportError:
            self.show_error_message(
                "Module PyMuPDF non installé",
                "Pour prévisualiser les PDF, installez: pip install PyMuPDF pillow"
            )
        except Exception as e:
            self.show_error_message("Erreur de prévisualisation PDF", str(e))
    
    def preview_text(self):
        """Prévisualiser un fichier texte"""
        try:
            with open(self.file['filepath'], 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)  # Limiter à 50KB
            
            text_widget = scrolledtext.ScrolledText(
                self.content_frame,
                wrap=tk.WORD,
                font=('Consolas', 10),
                bg='#ffffff',
                fg='#212529'
            )
            text_widget.pack(fill=tk.BOTH, expand=True)
            text_widget.insert('1.0', content)
            text_widget.config(state=tk.DISABLED)
            
        except Exception as e:
            self.show_error_message("Erreur de lecture", str(e))
    
    def preview_image(self):
        """Prévisualiser une image"""
        try:
            from PIL import Image, ImageTk
            
            # Ouvrir l'image
            img = Image.open(self.file['filepath'])
            
            # Redimensionner si trop grande
            max_size = (800, 600)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            photo = ImageTk.PhotoImage(img)
            
            # Afficher dans un label
            label = tk.Label(self.content_frame, image=photo, bg='white')
            label.image = photo  # Garder une référence
            label.pack(expand=True)
            
        except ImportError:
            self.show_error_message(
                "Module Pillow non installé",
                "Pour prévisualiser les images, installez: pip install pillow"
            )
        except Exception as e:
            self.show_error_message("Erreur de prévisualisation image", str(e))
    
    def preview_docx(self):
        """Prévisualiser un document Word"""
        try:
            from docx import Document
            
            doc = Document(self.file['filepath'])
            
            # Créer un widget texte scrollable
            text_widget = scrolledtext.ScrolledText(
                self.content_frame,
                wrap=tk.WORD,
                font=('Segoe UI', 11),
                bg='#ffffff',
                fg='#212529',
                padx=20,
                pady=20
            )
            text_widget.pack(fill=tk.BOTH, expand=True)
            
            # Extraire le texte
            for para in doc.paragraphs:
                text_widget.insert(tk.END, para.text + '\n\n')
            
            text_widget.config(state=tk.DISABLED)
            
        except ImportError:
            self.show_error_message(
                "Module python-docx non installé",
                "Pour prévisualiser les fichiers Word, installez: pip install python-docx"
            )
        except Exception as e:
            self.show_error_message("Erreur de prévisualisation Word", str(e))
    
    def preview_xlsx(self):
        """Prévisualiser un fichier Excel"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(self.file['filepath'], read_only=True, data_only=True)
            
            # Créer un notebook pour les feuilles
            notebook = ttk.Notebook(self.content_frame)
            notebook.pack(fill=tk.BOTH, expand=True)
            
            # Limiter à 3 feuilles
            for sheet_name in list(wb.sheetnames)[:3]:
                sheet = wb[sheet_name]
                
                # Frame pour cette feuille
                sheet_frame = tk.Frame(notebook, bg='white')
                notebook.add(sheet_frame, text=sheet_name)
                
                # Canvas avec scrollbars
                canvas = tk.Canvas(sheet_frame, bg='white')
                v_scrollbar = ttk.Scrollbar(sheet_frame, orient=tk.VERTICAL, command=canvas.yview)
                h_scrollbar = ttk.Scrollbar(sheet_frame, orient=tk.HORIZONTAL, command=canvas.xview)
                
                scrollable_frame = tk.Frame(canvas, bg='white')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor=tk.NW)
                canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
                
                canvas.grid(row=0, column=0, sticky='nsew')
                v_scrollbar.grid(row=0, column=1, sticky='ns')
                h_scrollbar.grid(row=1, column=0, sticky='ew')
                
                sheet_frame.grid_rowconfigure(0, weight=1)
                sheet_frame.grid_columnconfigure(0, weight=1)
                
                # Créer un tableau (limiter à 100 lignes pour les performances)
                max_rows = min(100, sheet.max_row)
                max_cols = min(20, sheet.max_column)
                
                for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, max_col=max_cols), 1):
                    for col_idx, cell in enumerate(row, 1):
                        cell_value = str(cell.value) if cell.value is not None else ""
                        
                        # Style différent pour l'en-tête
                        if row_idx == 1:
                            cell_label = tk.Label(
                                scrollable_frame,
                                text=cell_value,
                                font=('Segoe UI', 9, 'bold'),
                                bg='#e9ecef',
                                fg='#212529',
                                relief=tk.SOLID,
                                bd=1,
                                width=15,
                                anchor=tk.W,
                                padx=5,
                                pady=3
                            )
                        else:
                            cell_label = tk.Label(
                                scrollable_frame,
                                text=cell_value,
                                font=('Segoe UI', 9),
                                bg='white',
                                fg='#212529',
                                relief=tk.SOLID,
                                bd=1,
                                width=15,
                                anchor=tk.W,
                                padx=5,
                                pady=3
                            )
                        
                        cell_label.grid(row=row_idx-1, column=col_idx-1, sticky='ew')
            
            wb.close()
            
        except ImportError:
            self.show_error_message(
                "Module openpyxl non installé",
                "Pour prévisualiser les fichiers Excel, installez: pip install openpyxl"
            )
        except Exception as e:
            self.show_error_message("Erreur de prévisualisation Excel", str(e))
    
    def show_no_preview(self):
        """Afficher un message quand la prévisualisation n'est pas disponible"""
        message_frame = tk.Frame(self.content_frame, bg='white')
        message_frame.pack(expand=True)
        
        tk.Label(
            message_frame,
            text="⚠️",
            font=('Arial', 48),
            bg='white',
            fg='#ffc107'
        ).pack(pady=20)
        
        tk.Label(
            message_frame,
            text="Prévisualisation non disponible",
            font=('Segoe UI', 14, 'bold'),
            bg='white',
            fg='#212529'
        ).pack(pady=10)
        
        tk.Label(
            message_frame,
            text="Ce type de fichier ne peut pas être prévisualisé.\nUtilisez le bouton 'Ouvrir avec application' pour le consulter.",
            font=('Segoe UI', 10),
            bg='white',
            fg='#6c757d',
            justify=tk.CENTER
        ).pack(pady=10)
    
    def show_error_message(self, title: str, message: str):
        """Afficher un message d'erreur"""
        error_frame = tk.Frame(self.content_frame, bg='white')
        error_frame.pack(expand=True)
        
        tk.Label(
            error_frame,
            text="❌",
            font=('Arial', 48),
            bg='white',
            fg='#dc3545'
        ).pack(pady=20)
        
        tk.Label(
            error_frame,
            text=title,
            font=('Segoe UI', 14, 'bold'),
            bg='white',
            fg='#dc3545'
        ).pack(pady=10)
        
        tk.Label(
            error_frame,
            text=message,
            font=('Segoe UI', 10),
            bg='white',
            fg='#6c757d',
            justify=tk.CENTER,
            wraplength=600
        ).pack(pady=10)
    
    def open_with_default_app(self):
        """Ouvrir le fichier avec l'application par défaut"""
        success = self.file_handler.open_file(self.file['filepath'])
        if not success:
            messagebox.showerror("Erreur", "Impossible d'ouvrir le fichier")